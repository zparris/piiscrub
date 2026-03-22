"""Shared test fixtures for PIIScrub.

All document fixtures are generated programmatically — no binary files committed.
The analyzer fixture is session-scoped to avoid the 2-4s spaCy model load per test.
"""

from __future__ import annotations

import pytest
from pathlib import Path

PII_TEXT = (
    "Dear John Smith,\n"
    "Thank you for your enquiry. Your National Insurance number AB123456C has been "
    "recorded on our system. Please contact us at john.smith@example.co.uk or call "
    "07700 900123. Your NHS number is 943 476 5919.\n"
    "Your postcode SW1A 2AA and our reference GB29NWBK60161331926819.\n"
    "Kind regards,\nSarah Mitchell\nClient Services"
)


@pytest.fixture(scope="session")
def sample_txt(tmp_path_factory) -> Path:
    p = tmp_path_factory.mktemp("fixtures") / "sample.txt"
    p.write_text(PII_TEXT, encoding="utf-8")
    return p


@pytest.fixture(scope="session")
def sample_pdf(tmp_path_factory) -> Path:
    pytest.importorskip("fpdf", reason="fpdf2 required for PDF fixture generation")
    from fpdf import FPDF

    p = tmp_path_factory.mktemp("fixtures") / "sample.pdf"
    doc = FPDF()
    doc.add_page()
    doc.set_font("Helvetica", size=11)
    for line in PII_TEXT.splitlines():
        doc.cell(0, 8, line, ln=True)
    doc.output(str(p))
    return p


@pytest.fixture(scope="session")
def sample_docx(tmp_path_factory) -> Path:
    pytest.importorskip("docx", reason="python-docx required for DOCX fixture")
    from docx import Document

    p = tmp_path_factory.mktemp("fixtures") / "sample.docx"
    doc = Document()
    for line in PII_TEXT.splitlines():
        doc.add_paragraph(line)
    doc.save(str(p))
    return p


@pytest.fixture(scope="session")
def sample_xlsx(tmp_path_factory) -> Path:
    pytest.importorskip("openpyxl", reason="openpyxl required for XLSX fixture")
    import openpyxl

    p = tmp_path_factory.mktemp("fixtures") / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Name"
    ws["B1"] = "Email"
    ws["C1"] = "NI Number"
    ws["A2"] = "John Smith"
    ws["B2"] = "john.smith@example.co.uk"
    ws["C2"] = "AB123456C"
    wb.save(str(p))
    return p


@pytest.fixture(scope="session")
def sample_csv(tmp_path_factory) -> Path:
    p = tmp_path_factory.mktemp("fixtures") / "sample.csv"
    p.write_text(
        "name,email,ni_number\n"
        "John Smith,john.smith@example.co.uk,AB123456C\n"
        "Sarah Mitchell,sarah@example.co.uk,CD234567D\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture(scope="session")
def sample_eml(tmp_path_factory) -> Path:
    p = tmp_path_factory.mktemp("fixtures") / "sample.eml"
    p.write_text(
        "From: John Smith <john.smith@example.co.uk>\n"
        "To: Sarah Mitchell <sarah@example.co.uk>\n"
        "Subject: NI number query\n"
        "Content-Type: text/plain\n\n"
        "Dear Sarah,\nMy NI number is AB123456C. Please call 07700 900123.\n",
        encoding="utf-8",
    )
    return p


@pytest.fixture(scope="session")
def analyzer():
    """Session-scoped Presidio analyzer — loads spaCy model once for all tests."""
    import spacy
    if not spacy.util.is_package("en_core_web_lg"):
        pytest.skip("en_core_web_lg not installed — run: python -m spacy download en_core_web_lg")
    from piiscrub.detector import build_analyzer
    return build_analyzer()
